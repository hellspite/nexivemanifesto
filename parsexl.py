import openpyxl

"""
Funzione per il caricamento del file excel e la selezione del foglio di calcolo
INPUT nome del file
OUTPUT False se il file è sconosciuto oppure restituisce il foglio di calcolo
"""


def load_excel(file_name):
    try:
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        print("File non trovato")
        return False

    return wb


"""
Funzione per selezionare il foglio di calcolo dal documento excel
"""


def select_sheet(wb):
    ws = None
    while ws is None:
        sheet_num = int(input("Seleziona il numero del foglio di calcolo: "))
        wb.active = (sheet_num - 1)
        ws = wb.active
        if ws is None:
            print("Foglio di calcolo inesistente, provare di nuovo")
        else:
            if check_sheet(ws) is False:
                print(f'Hai selezionato il foglio: {ws.title}')
                print("Foglio di calcolo non valido, provare di nuovo")
                ws = None
            else:
                correct = None
                while correct != 's' and correct != 'n':
                    correct = input('Hai selezionato il foglio: ' + ws.title + ' è corretto? (s/n) ')
                    correct.lower()
                if correct == 'n':
                    ws = None

    return ws


"""
Funzione per creare un nuovo foglio excel che verrà poi caricato sul sito Nexive
"""


def create_empty_sheet():
    wb = openpyxl.Workbook()

    ws = wb.active

    ws['A1'] = "TITOLO"
    ws['B1'] = "COGNOME_RAGSOC"
    ws['C1'] = "NOME"
    ws['D1'] = "INFOUTILIXCONSEGNA"
    ws['E1'] = "VIA"
    ws['F1'] = "PIANOINTERNO"
    ws['G1'] = "CAP"
    ws['H1'] = "FRAZIONE"
    ws['I1'] = "COMUNE"
    ws['J1'] = "PROVINCIA"
    ws['K1'] = "IMPORTO_CONTRASSEGNO"
    ws['L1'] = "MODALITACONTRASSEGNO"
    ws['M1'] = "COLLI"
    ws['N1'] = "TAGLIA"
    ws['O1'] = "CELLULARE"
    ws['P1'] = "TELEFONO"
    ws['Q1'] = "EMAIL"
    ws['R1'] = "RIFERIMENTO_CLIENTE"
    ws['S1'] = "CENTRO_COSTO"
    ws['T1'] = "MITTENTE"
    ws['U1'] = "MITTENTE_NOME"
    ws['V1'] = "MITTENTE_VIA"
    ws['W1'] = "MITTENTE_CAP"
    ws['X1'] = "MITTENTE_COMUNE"
    ws['Y1'] = "MITTENTE_PROVINCIA"
    ws['Z1'] = "IMPORTOASSICURATA"
    ws['AA1'] = "PRODOTTO"
    ws['AB1'] = "SERVIZIO_RESI_EASY"

    return wb


"""
Funzione per contare le righe effettive da processare
"""


def count_rows(worksheet):
    col_a = worksheet['A']

    row_num: int = 0

    for i in col_a:
        if i.value is not None:
            row_num += 1

    return row_num


"""
Funzione che si occupa dell'elaborazione del contenuto
"""


def parse_content(quantity, content):
    content = content.lower()
    content = content.replace('maglietta io rompo black', 'B')
    content = content.replace('maglietta io rompo orange', 'O')
    content = content.replace('femmina', 'F')
    content = content.replace('maschio', 'M')
    content = content.upper()

    if quantity > 1:
        content = content + "   "
        content = content * quantity
        content = content.rstrip()

    return content


"""
Funzione per passare i dati dal primo foglio excel a quello definitivo per la spedizione Nexive
"""


def parse_xl(ws_in, wb_out):
    file_in = ws_in
    file_out = wb_out.active

    rows = count_rows(file_in) + 1

    # Numero d'ordine
    for i in range(3, rows):
        file_out['A' + str(i - 1)] = file_in['A' + str(i)].value

    # Email
    for i in range(3, rows):
        file_out['Q' + str(i - 1)] = file_in['D' + str(i)].value

    # Nome
    for i in range(3, rows):
        file_out['B' + str(i - 1)] = file_in['E' + str(i)].value

    # Indirizzo
    for i in range(3, rows):
        file_out['E' + str(i - 1)] = file_in['F' + str(i)].value

    # Presso
    for i in range(3, rows):
        file_out['C' + str(i - 1)] = file_in['G' + str(i)].value

    # Città
    for i in range(3, rows):
        file_out['I' + str(i - 1)] = file_in['H' + str(i)].value

    # CAP
    for i in range(3, rows):
        cap = file_in['I' + str(i)].value
        if cap is not None:
            cap = int(cap)
            cap = str(cap)
            cap_len = len(cap)
            if cap_len < 5:
                for c in range(cap_len, 5):
                    cap = '0' + cap
        else:
            cap = None

        file_out['G' + str(i - 1)] = cap

    # Provincia
    for i in range(3, rows):
        file_out['J' + str(i - 1)] = file_in['J' + str(i)].value

    # Telefono
    for i in range(3, rows):
        file_out['P' + str(i - 1)] = file_in['K' + str(i)].value

    # Taglia
    for i in range(3, rows):
        file_out['N' + str(i - 1)] = 'S'

    # Contenuto
    for i in range(3, rows):
        quantity = int(file_in['B' + str(i)].value)
        shirt = file_in['C' + str(i)].value

        shirt = parse_content(quantity, shirt)

        file_out['D' + str(i - 1)] = shirt

    return wb_out


"""
Funzione per verificare ordine su righe multiple
"""


def check_rows(wb):
    wb = wb
    ws = wb.active

    rows = count_rows(ws)

    prev_order = 0
    double_list = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=rows, min_col=1, max_col=4)):
        if row[0].value == prev_order:
            double_list.append((i - 1, i))

        prev_order = row[0].value

    if len(double_list) > 0:
        to_delete = []
        for o, d in double_list:
            o_index = o + 2
            d_index = d + 2

            ws['D' + str(o_index)] = ws['D' + str(o_index)].value + '   ' + ws['D' + str(d_index)].value

            to_delete.append(d_index)

        counter = 0
        for d in to_delete:
            ws.delete_rows(d - counter)

            counter += 1

    return wb


"""
Funzione per verificare validità foglio di calcolo
"""


def check_sheet(sheet):
    if sheet is None:
        return False

    if sheet.title[:6].lower() == 'ordini':
        return True
    elif sheet.title[:6].lower() == 'ordine':
        return True
    else:
        return False
