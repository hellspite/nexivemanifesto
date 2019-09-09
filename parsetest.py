import unittest
import parsexl
import openpyxl


class ParsexlTestCase(unittest.TestCase):
    def test_load_excel(self):
        file = "Manifesto.xlsx"
        wb = openpyxl.load_workbook(file)

        self.assertIsNotNone(wb)

    def test_load_excel_fail(self):
        file = "FileErrato.xlsx"
        wb_fail = None
        try:
            wb_fail = openpyxl.load_workbook(file)
        except FileNotFoundError:
            print("File non trovato!")

        self.assertIsNone(wb_fail)

    def test_create_empty_sheet(self):
        wb = parsexl.create_empty_sheet()
        ws = wb.active

        self.assertEqual(ws['A1'].value, 'TITOLO')

    def test_count_rows(self):
        wb = openpyxl.load_workbook('Manifesto.xlsx')
        wb.active = 1
        ws = wb.active

        rows = parsexl.count_rows(ws)

        self.assertEqual(rows, 11)

    def test_parse_shirts(self):
        quantity = 3
        content = "Maglietta io rompo black - Femmina / M"

        shirt = parsexl.parse_content(quantity, content)

        self.assertEqual(shirt, 'B - F / M   B - F / M   B - F / M')

    def test_parse_one_shirt(self):
        quantity = 1
        content = "Maglietta io rompo orange - Maschio / M"

        shirt = parsexl.parse_content(quantity, content)

        self.assertEqual(shirt, 'O - M / M')

    # TODO Scrivere test per il parse del contenuto

    # TODO Scrivere test per il check di ordini su righe multiple

    def test_check_sheet_none(self):
        result = parsexl.check_sheet(None)

        self.assertEqual(result, False)

    def test_check_sheet_correct(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ordini'

        result = parsexl.check_sheet(ws)

        self.assertEqual(result, True)

    def test_check_sheet_false(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Magazzino'

        result = parsexl.check_sheet(ws)

        self.assertEqual(result, False)



if __name__ == '__main__':
    unittest.main()
